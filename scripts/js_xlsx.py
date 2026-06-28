#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE_COLUMNS = [
    "id",
    "activity",
    "place",
    "arrival",
    "departure",
    "date",
    "tz",
    "lat",
    "lng",
    "note",
    "title",
    "photo_we",
    "photo_spots",
]

PHOTO_COLUMNS = {"photo_we", "photo_spots"}

# 照片名前缀。她在 Excel 里可以只填 IMG_001.jpg，
# 导出 JS 时会自动变成 }IMG_001.jpg。
PHOTO_PREFIX = "}"

JS_HEADER = "/* auto-generated — do not edit */\nexport const records = "


def extract_js_array(text: str) -> str:
    m = re.search(r"export\s+const\s+records\s*=", text)
    if not m:
        raise ValueError("没找到 `export const records =`。")

    start = text.find("[", m.end())
    if start == -1:
        raise ValueError("没找到 records 后面的 `[`。")

    depth = 0
    in_str = False
    quote = ""
    escaped = False

    for i in range(start, len(text)):
        ch = text[i]

        if in_str:
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == quote:
                in_str = False
            continue

        if ch in ('"', "'"):
            in_str = True
            quote = ch
            continue

        if ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                return text[start:i + 1]

    raise ValueError("records 数组没有正确结束，少了 `]`。")


def remove_trailing_commas(s: str) -> str:
    return re.sub(r",\s*(?=[}\]])", "", s)


def load_records_from_js(js_path: Path) -> list[dict[str, Any]]:
    text = js_path.read_text(encoding="utf-8")
    arr_text = extract_js_array(text)
    arr_text = remove_trailing_commas(arr_text)
    data = json.loads(arr_text)
    if not isinstance(data, list):
        raise ValueError("records 不是数组。")
    return data


def ordered_columns(records: list[dict[str, Any]]) -> list[str]:
    seen = set(BASE_COLUMNS)
    extras = []

    for rec in records:
        for key in rec.keys():
            if key not in seen:
                seen.add(key)
                extras.append(key)

    return BASE_COLUMNS + extras


def js_value_to_xlsx(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, list):
        if all(not isinstance(x, (dict, list)) for x in value):
            return "\n".join(str(x) for x in value)
        return json.dumps(value, ensure_ascii=False)

    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)

    return value


def js_to_xlsx(js_path: Path, xlsx_path: Path) -> None:
    records = load_records_from_js(js_path)
    cols = ordered_columns(records)

    wb = Workbook()
    ws = wb.active
    ws.title = "records"

    ws.append(cols)

    for rec in records:
        ws.append([js_value_to_xlsx(rec.get(col)) for col in cols])

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(cols, start=1):
        letter = get_column_letter(col_idx)

        width = 16
        if col_name in {"arrival", "departure"}:
            width = 26
        elif col_name in {"place", "note", "title", "photo_we", "photo_spots"}:
            width = 28
        elif col_name in {"lat", "lng"}:
            width = 14

        ws.column_dimensions[letter].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def is_empty(v: Any) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def to_int(v: Any) -> int | None:
    if is_empty(v):
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    return int(str(v).strip())


def to_float(v: Any) -> float | None:
    if is_empty(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return float(str(v).strip())


def to_text(v: Any, col: str) -> str | None:
    if is_empty(v):
        return None

    if isinstance(v, datetime):
        if col in {"arrival", "departure"}:
            return v.strftime("%Y-%m-%dT%H:%M:%S+0800")
        if col == "date":
            return v.strftime("%Y-%m-%d")

    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")

    return str(v).strip()


def normalize_photo_name(name: str) -> str:
    name = name.strip()

    if not name:
        return name

    if name.startswith((PHOTO_PREFIX, "http://", "https://", "/")):
        return name

    return PHOTO_PREFIX + name


def parse_photo_cell(v: Any) -> list[str] | None:
    """
    Excel 照片列规则：
    1. 空白 = null
    2. 一张照片：直接填 IMG_001.jpg
    3. 多张照片：一行一张
    """
    if is_empty(v):
        return None

    text = str(v).strip()

    if not text:
        return None

    if text.startswith("["):
        try:
            arr = json.loads(remove_trailing_commas(text))
            if isinstance(arr, list):
                return [normalize_photo_name(str(x)) for x in arr if not is_empty(x)] or None
        except Exception:
            pass

    parts = [p.strip() for p in re.split(r"[\n,，;；]+", text) if p.strip()]
    return [normalize_photo_name(p) for p in parts] or None


def parse_general_json_like(v: Any) -> Any:
    if is_empty(v):
        return None

    if not isinstance(v, str):
        return v

    text = v.strip()

    if text in {"null", "None", "none"}:
        return None

    if text in {"true", "false"}:
        return text == "true"

    if text.startswith("[") or text.startswith("{"):
        try:
            return json.loads(remove_trailing_commas(text))
        except Exception:
            return text

    return text


def xlsx_value_to_js(col: str, v: Any) -> Any:
    if col == "id":
        return to_int(v)

    if col in {"lat", "lng"}:
        return to_float(v)

    if col in PHOTO_COLUMNS:
        return parse_photo_cell(v)

    if col in {
        "activity",
        "place",
        "arrival",
        "departure",
        "date",
        "tz",
        "note",
        "title",
    }:
        return to_text(v, col)

    return parse_general_json_like(v)


def xlsx_to_records(xlsx_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    headers = [
        str(c.value).strip() if c.value is not None else ""
        for c in ws[1]
    ]

    if not any(headers):
        raise ValueError("xlsx 第一行必须是表头，比如 id、activity、place。")

    records = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(is_empty(v) for v in row):
            continue

        rec = {}

        for col, v in zip(headers, row):
            if not col:
                continue
            rec[col] = xlsx_value_to_js(col, v)

        records.append(rec)

    used_ids = [
        r.get("id")
        for r in records
        if isinstance(r.get("id"), int)
    ]

    next_id = max(used_ids) + 1 if used_ids else 1

    for rec in records:
        if rec.get("id") is None:
            rec["id"] = next_id
            next_id += 1

    return records


def dump_js(records: list[dict[str, Any]], js_path: Path) -> None:
    cols = ordered_columns(records)
    ordered_records = []

    for rec in records:
        ordered = {
            col: rec.get(col)
            for col in cols
            if col in rec
        }
        ordered_records.append(ordered)

    body = json.dumps(ordered_records, ensure_ascii=False, indent=2)

    js_path.parent.mkdir(parents=True, exist_ok=True)
    js_path.write_text(JS_HEADER + body + "\n", encoding="utf-8")


def xlsx_to_js(xlsx_path: Path, js_path: Path) -> None:
    records = xlsx_to_records(xlsx_path)
    dump_js(records, js_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="records.js 和 records.xlsx 互相转换")
    parser.add_argument("mode", choices=["js2xlsx", "xlsx2js"])
    parser.add_argument("input")
    parser.add_argument("output")

    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if args.mode == "js2xlsx":
        js_to_xlsx(input_path, output_path)
    else:
        xlsx_to_js(input_path, output_path)

    print(f"转换完成：{output_path}")


if __name__ == "__main__":
    main()