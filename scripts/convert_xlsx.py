"""
Convert 记录_together.xlsx → js/data/records.js as an ES module.

Filters records to dates present in SHAPE_INPUT (mirrors config.js logic).
Re-run this whenever the xlsx file or config.js SHAPE_INPUT changes.

Usage:
  python3 -m venv /tmp/xlsxenv
  source /tmp/xlsxenv/bin/activate
  pip install openpyxl
  python3 scripts/convert_xlsx.py
"""
import openpyxl, json, os
from datetime import datetime, timedelta

SRC = os.path.join(os.path.dirname(__file__), "..", "记录_together.xlsx")
DST = os.path.join(os.path.dirname(__file__), "..", "js", "data", "records.js")

# ── config mirror (keep in sync with config.js) ──────────────────

YEAR = 2026


def range_md(start_md, end_md, year):
    """Generate "M.D" strings from start to end (inclusive), like JS rangeMD()."""
    def parse(md):
        m, d = md.split(".")
        return datetime(year, int(m), int(d))

    result = []
    cur = parse(start_md)
    end = parse(end_md)
    while cur <= end:
        result.append(f"{cur.month}.{cur.day}")
        cur += timedelta(days=1)
    return result


SHAPE_INPUT_MD = set()
for mds in [
    ["5.5", "5.19", "6.10"],                                              # top
    ["2.17", "4.6", "5.1"],                                               # middle
    ["2.16", "2.22", "3.6", "3.13", "3.20", "5.14", "5.22", "5.27"],     # bottom
    [                                                                      # full
        "2.21", "2.26",
        "3.7", "3.8", "3.14", "3.15", "3.16", "3.21",
        "4.4", "4.5", "4.18", "4.19", "4.20",
        "5.2", "5.3", "5.4",
        "5.15", "5.16", "5.17", "5.18",
        "5.23", "5.24",
        *range_md("5.28", "6.9", YEAR),
        *range_md("6.18", "6.21", YEAR),
    ],
]:
    SHAPE_INPUT_MD.update(mds)


def to_date_str(md, year):
    m, d = md.split(".")
    return f"{year}-{int(m):02d}-{int(d):02d}"


SHAPE_DATE_SET = {to_date_str(md, YEAR) for md in SHAPE_INPUT_MD}

# ── read & filter ────────────────────────────────────────────────

wb = openpyxl.load_workbook(SRC)
ws = wb.active

records = []
skipped = 0
rid = 0

for r in range(2, ws.max_row + 1):
    activity = ws.cell(r, 1).value
    place = ws.cell(r, 2).value
    arrival = ws.cell(r, 3).value
    departure = ws.cell(r, 4).value
    tz = ws.cell(r, 5).value
    lat = ws.cell(r, 6).value
    lng = ws.cell(r, 7).value
    note = ws.cell(r, 8).value

    if not activity or not place or lat is None or lng is None:
        continue

    def norm_dt(val):
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%dT%H:%M:%S%z")
        return str(val)

    date_str = norm_dt(arrival)[:10] if arrival else None

    # filter: only keep records whose date is in SHAPE_INPUT
    if date_str not in SHAPE_DATE_SET:
        skipped += 1
        continue

    rid += 1
    records.append({
        "id": rid,
        "activity": str(activity).strip(),
        "place": str(place).strip(),
        "arrival": norm_dt(arrival),
        "departure": norm_dt(departure),
        "date": date_str,
        "tz": str(tz).strip() if tz else "Asia/Shanghai",
        "lat": round(float(lat), 7),
        "lng": round(float(lng), 7),
        "note": str(note).strip() if note else None
    })

# ── write ES module ─────────────────────────────────────────────

js = f"/* auto-generated — do not edit */\nexport const records = {json.dumps(records, ensure_ascii=False, indent=2)};\n"

with open(DST, "w") as f:
    f.write(js)

print(f"✓ Written {len(records)} records to {DST}  (skipped {skipped} outside SHAPE_INPUT dates)")
