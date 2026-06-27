"""
Convert 交通_together.xlsx → js/data/traffic.js.

Matches traffic to gaps between consecutive records (sorted by time).
  origin = the record before the gap
  dest   = the record after the gap
Traffic that doesn't fall in any gap is skipped — this naturally filters
out trips to/from deleted places and redundant round-trips.

Usage:
  source /tmp/xlsxenv/bin/activate
  python3 scripts/convert_traffic.py
"""
import json, os, re
from datetime import datetime, timedelta
import openpyxl

SRC_RECORDS_JS = os.path.join(os.path.dirname(__file__), "..", "js", "data", "records.js")
SRC_TRAFFIC = os.path.join(os.path.dirname(__file__), "..", "交通_together.xlsx")
DST = os.path.join(os.path.dirname(__file__), "..", "js", "data", "traffic.js")

TOLERANCE = timedelta(minutes=2)


def parse_dt(s):
    """Parse ISO 8601 string like '2026-06-23T02:24:43+0800'."""
    s_clean = s[:19]
    return datetime.strptime(s_clean, "%Y-%m-%dT%H:%M:%S")


# ── load records from hand-edited records.js ────────────────────

with open(SRC_RECORDS_JS, "r") as f:
    js_text = f.read()

m = re.search(r"export const records = (\[.*?\]);", js_text, re.DOTALL)
if not m:
    print("ERROR: could not parse records.js")
    exit(1)
arr_text = re.sub(r",\s*\]", "]", m.group(1))
records_data = json.loads(arr_text)

# build sorted list of records
recs = []
for r in records_data:
    if not r.get("arrival") or not r.get("departure"):
        continue
    if r.get("lat") is None or r.get("lng") is None:
        continue
    recs.append({
        "id": r["id"],
        "place": r["place"],
        "arrival": parse_dt(r["arrival"]),
        "departure": parse_dt(r["departure"]),
        "lat": r["lat"],
        "lng": r["lng"],
        "date": r["date"],
    })

recs.sort(key=lambda r: r["arrival"])

# build gap index: each gap = (rec[i].departure, rec[i+1].arrival, rec[i], rec[i+1])
gaps = []
for i in range(len(recs) - 1):
    gaps.append({
        "start": recs[i]["departure"],
        "end": recs[i + 1]["arrival"],
        "origin_rec": recs[i],
        "dest_rec": recs[i + 1],
    })

record_dates = {r["date"] for r in recs}


def find_gap(t_from, t_to):
    """Find the gap that contains this traffic time window (with tolerance)."""
    for g in gaps:
        if g["start"] - TOLERANCE <= t_from and t_to <= g["end"] + TOLERANCE:
            return g
    return None


# ── read & convert traffic ──────────────────────────────────────

wb_tr = openpyxl.load_workbook(SRC_TRAFFIC)
ws_tr = wb_tr.active

traffic = []
skipped_date = 0
skipped_gap = 0

for r in range(2, ws_tr.max_row + 1):
    typ = ws_tr.cell(r, 1).value
    from_time = ws_tr.cell(r, 2).value
    to_time = ws_tr.cell(r, 3).value
    tz = ws_tr.cell(r, 4).value
    dur = ws_tr.cell(r, 5).value
    origin_name = ws_tr.cell(r, 6).value
    dest_name = ws_tr.cell(r, 7).value

    if not typ or not from_time or not to_time:
        continue

    date_str = from_time[:10]
    if date_str not in record_dates:
        skipped_date += 1
        continue

    f_dt = parse_dt(from_time)
    t_dt = parse_dt(to_time)
    gap = find_gap(f_dt, t_dt)
    if not gap:
        skipped_gap += 1
        continue

    o = gap["origin_rec"]
    d = gap["dest_rec"]

    traffic.append({
        "type": str(typ).strip(),
        "from_time": from_time,
        "to_time": to_time,
        "tz": str(tz).strip() if tz else "Asia/Shanghai",
        "duration_sec": int(dur) if dur else None,
        "origin": str(origin_name).strip() if origin_name else o["place"],
        "origin_lat": o["lat"],
        "origin_lng": o["lng"],
        "origin_record_id": o["id"],
        "dest": str(dest_name).strip() if dest_name else d["place"],
        "dest_lat": d["lat"],
        "dest_lng": d["lng"],
        "dest_record_id": d["id"],
        "date": date_str,
    })

# ── write ES module ─────────────────────────────────────────────

js = (
    "/* auto-generated — do not edit */\n"
    f"export const traffic = {json.dumps(traffic, ensure_ascii=False, indent=2)};\n"
)

with open(DST, "w") as f:
    f.write(js)

print(f"Written {len(traffic)} traffic records to {os.path.basename(DST)}")
print(f"  skipped {skipped_date} outside record dates, {skipped_gap} not in any gap")
